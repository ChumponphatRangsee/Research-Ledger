from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from uuid import UUID

import pytest
from openpyxl import Workbook, load_workbook

from app.services.portfolio_import import cli as import_cli
from app.services.portfolio_import import (
    REQUIRED_SHEET_TITLES,
    FeeUnit,
    SupabaseTransactionImportRepository,
    WorkbookStructureError,
    apply_existing_fingerprint_deduplication,
    build_import_plan,
)
from app.services.portfolio_import.google_sheets import TRANSACTION_HEADERS

SPREADSHEET_ID = "current-15-tab-sheet"
USER_ID = UUID("11111111-1111-4111-8111-111111111111")


def _row(
    source_id: str,
    transaction_date: str,
    account: str,
    asset: str,
    asset_class: str,
    action: str,
    quantity: float,
    price: float | str,
    fee: float,
    fee_unit: str,
    currency: str,
    fx_rate: float,
    notes: str = "",
) -> list[object]:
    return [
        source_id,
        transaction_date,
        account,
        asset,
        asset_class,
        action,
        quantity,
        price,
        fee,
        fee_unit,
        currency,
        fx_rate,
        "=G5",
        "=G5*H5*L5",
        "=0",
        "=0",
        "=0",
        "=0",
        "=0",
        "=0",
        "OK",
        notes,
    ]


def _write_workbook(
    path: Path,
    *,
    rows: list[list[object]] | None = None,
    holdings: list[tuple[str, str, Decimal]] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.active.title = REQUIRED_SHEET_TITLES[0]
    for title in REQUIRED_SHEET_TITLES[1:]:
        workbook.create_sheet(title)

    transactions = workbook["Transactions"]
    for column, header in enumerate(TRANSACTION_HEADERS, start=1):
        transactions.cell(4, column, header)

    source_rows = rows or [
        _row(
            "SRC-BTC",
            "2026-02-07",
            "Best",
            "btc",
            "Crypto",
            "buy",
            0.0023,
            68590.01,
            0.0000023,
            "Asset Units",
            "USDT",
            33.44,
        ),
        _row(
            "SRC-CRWD-BUY",
            "2026-03-09",
            "Best",
            "CRWD",
            "Stock",
            "BUY",
            0.7354496,
            425.59,
            0,
            "Quote Currency",
            "USD",
            31.74,
        ),
        _row(
            "SRC-MSFT-1",
            "2026-05-10",
            "Best",
            "MSFT",
            "Stock",
            "BUY",
            0.434,
            180.71,
            0.27,
            "Quote Currency",
            "USD",
            32.27,
        ),
        _row(
            "SRC-CRWD-SELL-1",
            "2026-05-26",
            "Best",
            "CRWD",
            "Stock",
            "SELL",
            0.2694791,
            671.66,
            0.27,
            "Quote Currency",
            "USD",
            32.629,
        ),
        _row(
            "SRC-CRWD-SELL-2",
            "2026-05-29",
            "Best",
            "CRWD",
            "Stock",
            "SELL",
            0.4659704,
            675.86,
            0.47,
            "Quote Currency",
            "USD",
            32.55,
        ),
        _row(
            "SRC-MSFT-2",
            "2026-07-02",
            "Best",
            "MSFT",
            "Stock",
            "BUY",
            0.796852,
            385.26,
            0,
            "Quote Currency",
            "USD",
            33.19,
        ),
    ]
    for row_number, values in enumerate(source_rows, start=5):
        for column, value in enumerate(values, start=1):
            transactions.cell(row_number, column, value)

    holdings_sheet = workbook["Holdings"]
    for column, header in enumerate(
        ("Account", "Asset", "Target Bucket", "Quantity"), start=1
    ):
        holdings_sheet.cell(4, column, header)
    holding_rows = holdings or [
        ("Best", "BTC", Decimal("0.0022977")),
        ("Best", "CRWD", Decimal(0)),
        ("Best", "MSFT", Decimal("1.230852")),
    ]
    for row_number, (account, symbol, quantity) in enumerate(holding_rows, start=5):
        holdings_sheet.cell(row_number, 1, account)
        holdings_sheet.cell(row_number, 2, symbol)
        holdings_sheet.cell(row_number, 3, symbol)
        holdings_sheet.cell(row_number, 4, float(quantity))

    workbook.save(path)
    workbook.close()
    return path


def test_current_workbook_builds_review_only_plan(tmp_path: Path):
    plan = build_import_plan(
        _write_workbook(tmp_path / "portfolio.xlsx"),
        spreadsheet_id=SPREADSHEET_ID,
    )

    assert plan.rows_read == 6
    assert plan.issues == []
    assert len(plan.transactions) == 6
    assert plan.positions[("Best", "BTC")] == Decimal("0.0022977")
    assert plan.positions[("Best", "CRWD")] == 0
    assert plan.positions[("Best", "MSFT")] == Decimal("1.230852")
    assert all(check["passed"] for check in plan.checks.values())
    btc = next(row for row in plan.transactions if row.symbol == "BTC")
    assert btc.asset_currency == "USD"
    assert btc.currency == "USDT"
    assert btc.fee_unit == FeeUnit.ASSET_UNITS
    assert btc.fee_amount == Decimal("0.0000023")
    assert btc.raw_source_data["sheet"] == "Transactions"
    assert btc.raw_source_data["row_number"] == 5
    assert btc.source_metadata["formula_derived_summary_columns_ignored"] is True
    assert plan.report()["counts"]["asset_unit_fee_rows"] == 1


def test_formula_derived_transaction_price_is_isolated(tmp_path: Path):
    path = _write_workbook(tmp_path / "formula-price.xlsx")
    workbook = load_workbook(path)
    workbook["Transactions"]["H5"] = "=100+1"
    workbook.save(path)
    workbook.close()

    plan = build_import_plan(path, spreadsheet_id=SPREADSHEET_ID)

    assert any(issue.error_code == "FORMULA_DERIVED_PRICE" for issue in plan.issues)
    assert all(row.source_identifier != "SRC-BTC" for row in plan.transactions)
    assert plan.checks["formula_derived_transaction_prices"]["passed"] is False


def test_duplicate_source_fingerprint_is_not_staged_twice(tmp_path: Path):
    duplicated = _row(
        "DUPLICATE",
        "2026-01-01",
        "Best",
        "MSFT",
        "Stock",
        "BUY",
        1,
        100,
        0,
        "Quote Currency",
        "USD",
        33,
    )
    path = _write_workbook(
        tmp_path / "duplicate.xlsx",
        rows=[duplicated, duplicated],
        holdings=[("Best", "MSFT", Decimal(1))],
    )

    plan = build_import_plan(path, spreadsheet_id=SPREADSHEET_ID)

    assert len(plan.transactions) == 1
    assert any(
        issue.error_code == "DUPLICATE_SOURCE_FINGERPRINT" for issue in plan.issues
    )


def test_existing_source_fingerprint_is_removed_before_staging(tmp_path: Path):
    plan = build_import_plan(
        _write_workbook(tmp_path / "existing.xlsx"),
        spreadsheet_id=SPREADSHEET_ID,
    )
    existing = plan.transactions[0].source_fingerprint

    apply_existing_fingerprint_deduplication(plan, {existing})

    assert all(row.source_fingerprint != existing for row in plan.transactions)
    assert plan.report()["counts"]["error_codes"]["DUPLICATE_SOURCE_FINGERPRINT"] == 1


def test_sell_that_would_create_negative_position_becomes_error(tmp_path: Path):
    sell = _row(
        "NEGATIVE",
        "2026-01-01",
        "Best",
        "MSFT",
        "Stock",
        "SELL",
        1,
        100,
        0,
        "Quote Currency",
        "USD",
        33,
    )
    plan = build_import_plan(
        _write_workbook(tmp_path / "negative.xlsx", rows=[sell], holdings=[]),
        spreadsheet_id=SPREADSHEET_ID,
    )

    assert plan.transactions == []
    assert any(issue.error_code == "NEGATIVE_POSITION" for issue in plan.issues)
    assert plan.checks["non_negative_positions"]["passed"] is False


def test_obsolete_seven_tab_workbook_is_rejected(tmp_path: Path):
    path = _write_workbook(tmp_path / "old.xlsx")
    workbook = load_workbook(path)
    for title in REQUIRED_SHEET_TITLES[7:]:
        del workbook[title]
    workbook.save(path)
    workbook.close()

    with pytest.raises(WorkbookStructureError, match="current 15-tab"):
        build_import_plan(path, spreadsheet_id=SPREADSHEET_ID)


class RecordingQuery:
    def __init__(self, client: RecordingClient, table: str):
        self.client = client
        self.table = table
        self.operation: str | None = None
        self.payload = None
        self.filters: list[tuple[str, object]] = []
        self.conflict: str | None = None

    def select(self, columns: str):
        self.operation = "select"
        self.payload = columns
        return self

    def insert(self, payload):
        self.operation = "insert"
        self.payload = payload
        return self

    def upsert(self, payload, *, on_conflict: str):
        self.operation = "upsert"
        self.payload = payload
        self.conflict = on_conflict
        return self

    def update(self, payload):
        self.operation = "update"
        self.payload = payload
        return self

    def eq(self, column: str, value):
        self.filters.append((column, value))
        return self

    def in_(self, column: str, values):
        self.filters.append((column, tuple(values)))
        return self

    def order(self, column: str, desc: bool = False):
        self.filters.append(("order", column, desc))
        return self

    def limit(self, count: int):
        self.filters.append(("limit", count))
        return self

    def execute(self):
        self.client.queries.append(self)
        if self.operation == "select":
            return SimpleNamespace(data=self.client.select_data.get(self.table, []))
        generated_id = {
            "transaction_import_batches": "batch-id",
            "investment_accounts": "account-id",
            "assets": "asset-id",
        }.get(self.table, "row-id")
        return SimpleNamespace(data=[{"id": generated_id}])


class RecordingClient:
    def __init__(self):
        self.queries: list[RecordingQuery] = []
        self.select_data: dict[str, list[dict[str, object]]] = {}

    def table(self, table: str):
        return RecordingQuery(self, table)


class RecordingAuthAdmin:
    def __init__(self, pages: dict[int, list[object]]):
        self.pages = pages
        self.calls: list[tuple[int, int]] = []

    def list_users(self, *, page: int, per_page: int):
        self.calls.append((page, per_page))
        return self.pages.get(page, [])


class RecordingAuthClient:
    def __init__(self, pages: dict[int, list[object]]):
        self.auth = SimpleNamespace(admin=RecordingAuthAdmin(pages))


def test_cli_requires_explicit_dev_owner_bypass(monkeypatch):
    monkeypatch.delenv("SUPABASE_ACCESS_TOKEN", raising=False)
    monkeypatch.delenv("ALLOW_DEV_OWNER_BYPASS", raising=False)
    monkeypatch.delenv("DEV_IMPORT_USER_EMAIL", raising=False)

    with pytest.raises(SystemExit, match="SUPABASE_ACCESS_TOKEN is required"):
        import_cli._resolve_staging_user_id()


def test_cli_keeps_verified_jwt_as_primary_owner_source(monkeypatch):
    monkeypatch.setenv("SUPABASE_ACCESS_TOKEN", "jwt-token")
    monkeypatch.setenv("ALLOW_DEV_OWNER_BYPASS", "true")
    monkeypatch.setenv("DEV_IMPORT_USER_EMAIL", "dev@example.com")
    monkeypatch.setattr(
        import_cli,
        "verify_supabase_jwt",
        lambda token: SimpleNamespace(id=USER_ID),
    )

    def fail_lookup(_: str):
        raise AssertionError("dev owner bypass should not run when JWT is present")

    monkeypatch.setattr(import_cli, "_lookup_dev_owner_user_id", fail_lookup)

    assert import_cli._resolve_staging_user_id() == USER_ID


def test_cli_dev_owner_bypass_resolves_existing_auth_user(monkeypatch):
    owner_id = UUID("22222222-2222-4222-8222-222222222222")
    client = RecordingAuthClient(
        {
            1: [
                SimpleNamespace(id=str(USER_ID), email="other@example.com"),
                {"id": str(owner_id), "email": "Owner@Example.com"},
            ],
        }
    )
    monkeypatch.setenv("ALLOW_DEV_OWNER_BYPASS", "true")
    monkeypatch.setenv("DEV_IMPORT_USER_EMAIL", "owner@example.com")
    monkeypatch.delenv("SUPABASE_ACCESS_TOKEN", raising=False)
    monkeypatch.setattr(import_cli, "get_supabase_client", lambda: client)

    assert import_cli._resolve_staging_user_id() == owner_id
    assert client.auth.admin.calls == [(1, 1000)]


def test_repository_scopes_dedup_reads_to_verified_owner():
    client = RecordingClient()
    client.select_data = {
        "transaction_drafts": [{"source_fingerprint": "draft-fp"}],
        "transactions": [{"source_fingerprint": "confirmed-fp"}],
    }
    repository = SupabaseTransactionImportRepository(client)

    actual = repository.existing_source_fingerprints(
        user_id=USER_ID,
        source_fingerprints={"draft-fp", "confirmed-fp"},
    )

    assert actual == {"draft-fp", "confirmed-fp"}
    assert {query.table for query in client.queries} == {
        "transaction_drafts",
        "transactions",
    }
    assert all(("user_id", str(USER_ID)) in query.filters for query in client.queries)


def test_repository_lists_import_batches_for_verified_owner():
    client = RecordingClient()
    client.select_data = {
        "transaction_import_batches": [{"id": "batch-id", "user_id": str(USER_ID)}],
    }

    batches = SupabaseTransactionImportRepository(client).list_batches(
        user_id=USER_ID,
        limit=25,
    )

    assert batches == [{"id": "batch-id", "user_id": str(USER_ID)}]
    query = client.queries[0]
    assert query.table == "transaction_import_batches"
    assert ("user_id", str(USER_ID)) in query.filters
    assert ("order", "created_at", True) in query.filters
    assert ("limit", 25) in query.filters


def test_repository_lists_import_errors_for_verified_owner_and_batch():
    client = RecordingClient()
    client.select_data = {
        "transaction_import_errors": [{"id": "error-id", "user_id": str(USER_ID)}],
    }

    errors = SupabaseTransactionImportRepository(client).list_errors(
        user_id=USER_ID,
        import_batch_id=UUID("33333333-3333-4333-8333-333333333333"),
        limit=40,
    )

    assert errors == [{"id": "error-id", "user_id": str(USER_ID)}]
    query = client.queries[0]
    assert query.table == "transaction_import_errors"
    assert ("user_id", str(USER_ID)) in query.filters
    assert (
        "import_batch_id",
        "33333333-3333-4333-8333-333333333333",
    ) in query.filters
    assert ("order", "created_at", True) in query.filters
    assert ("limit", 40) in query.filters


def test_repository_stages_drafts_and_never_confirms_transactions(tmp_path: Path):
    plan = build_import_plan(
        _write_workbook(tmp_path / "stage.xlsx"),
        spreadsheet_id=SPREADSHEET_ID,
    )
    client = RecordingClient()

    SupabaseTransactionImportRepository(client).stage(plan, user_id=USER_ID)

    inserted_tables = [
        query.table for query in client.queries if query.operation == "insert"
    ]
    assert "transaction_import_batches" in inserted_tables
    assert "transaction_drafts" in inserted_tables
    assert "transactions" not in inserted_tables
    drafts = next(
        query.payload
        for query in client.queries
        if query.table == "transaction_drafts" and query.operation == "insert"
    )
    assert all(row["user_id"] == str(USER_ID) for row in drafts)
    btc = next(row for row in drafts if row["source_identifier"] == "SRC-BTC")
    assert btc["fee_unit"] == "ASSET_UNITS"
    assert btc["fee_amount"] == "0.0000023"
    batch_updates = [
        query
        for query in client.queries
        if query.table == "transaction_import_batches" and query.operation == "update"
    ]
    assert all(("user_id", str(USER_ID)) in query.filters for query in batch_updates)
    assert (
        batch_updates[-1].payload["source_metadata"]["confirmed_transactions_written"]
        == 0
    )


def test_repository_does_not_overwrite_existing_account_or_asset(tmp_path: Path):
    plan = build_import_plan(
        _write_workbook(tmp_path / "existing-identities.xlsx"),
        spreadsheet_id=SPREADSHEET_ID,
    )
    client = RecordingClient()
    client.select_data = {
        "investment_accounts": [{"id": "existing-account"}],
        "assets": [{"id": "existing-asset"}],
    }

    SupabaseTransactionImportRepository(client).stage(plan, user_id=USER_ID)

    identity_inserts = [
        query
        for query in client.queries
        if query.operation == "insert"
        and query.table in {"investment_accounts", "assets"}
    ]
    assert identity_inserts == []
    drafts = next(
        query.payload
        for query in client.queries
        if query.table == "transaction_drafts" and query.operation == "insert"
    )
    assert all(row["investment_account_id"] == "existing-account" for row in drafts)
    assert all(row["asset_id"] == "existing-asset" for row in drafts)
