"""Read the approved 15-tab Google Sheets XLSX export without trusting formulas."""

from __future__ import annotations

import hashlib
from datetime import date, datetime
from decimal import Decimal
from itertools import zip_longest
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.services.portfolio_import.models import (
    WorkbookSnapshot,
    WorkbookTransactionRow,
)

REQUIRED_SHEET_TITLES = (
    "Instructions",
    "Dashboard",
    "Transactions",
    "Price Data",
    "Holdings",
    "Target Allocation",
    "DCA Planner",
    "Risk Dashboard",
    "Thesis Tracker",
    "Performance",
    "Checks",
    "Lists",
    "AI Portfolio Diagnosis",
    "Recommended Actions",
    "AI Integration Config",
)

TRANSACTION_HEADER_ROW = 4
TRANSACTION_FIRST_ROW = 5
TRANSACTION_HEADERS = (
    "Source ID",
    "Date",
    "Account",
    "Asset",
    "Asset Class",
    "Action",
    "Quantity",
    "Price",
    "Fee",
    "Fee Unit",
    "Currency",
    "FX Rate",
    "Net Quantity",
    "Gross Value THB",
    "Net Cash Flow THB",
    "Cost Basis Change THB",
    "Realized P&L THB",
    "Running Qty",
    "Running Cost Basis THB",
    "Avg Cost / Unit THB",
    "Data Check",
    "Notes",
)


class WorkbookStructureError(ValueError):
    """Raised when the export is not the approved current workbook shape."""


def _json_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


class GoogleSheetsWorkbookReader:
    """Load authoritative inputs and cached reconciliation values from XLSX."""

    def read(self, path: str | Path) -> WorkbookSnapshot:
        workbook_path = Path(path)
        workbook_bytes = workbook_path.read_bytes()
        source_fingerprint = hashlib.sha256(workbook_bytes).hexdigest()

        formulas_book = load_workbook(workbook_path, data_only=False, read_only=True)
        values_book = load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            self._validate_structure(formulas_book.sheetnames)
            transaction_rows = self._read_transactions(formulas_book, values_book)
            holdings = self._read_holdings(values_book)
        finally:
            formulas_book.close()
            values_book.close()

        return WorkbookSnapshot(
            source_filename=workbook_path.name,
            source_fingerprint=source_fingerprint,
            sheet_titles=tuple(REQUIRED_SHEET_TITLES),
            transaction_rows=tuple(transaction_rows),
            holdings=holdings,
        )

    @staticmethod
    def _validate_structure(sheet_titles: list[str]) -> None:
        if tuple(sheet_titles) != REQUIRED_SHEET_TITLES:
            missing = sorted(set(REQUIRED_SHEET_TITLES) - set(sheet_titles))
            unexpected = sorted(set(sheet_titles) - set(REQUIRED_SHEET_TITLES))
            raise WorkbookStructureError(
                "Expected the current 15-tab portfolio workbook export; "
                f"found {len(sheet_titles)} tabs (missing={missing}, unexpected={unexpected})"
            )

    @staticmethod
    def _read_transactions(
        formulas_book: Any, values_book: Any
    ) -> list[WorkbookTransactionRow]:
        formula_sheet = formulas_book["Transactions"]
        value_sheet = values_book["Transactions"]
        observed_headers = tuple(
            formula_sheet.cell(TRANSACTION_HEADER_ROW, column).value
            for column in range(1, len(TRANSACTION_HEADERS) + 1)
        )
        if observed_headers != TRANSACTION_HEADERS:
            raise WorkbookStructureError(
                "Transactions headers do not match the approved 15-tab export"
            )

        rows: list[WorkbookTransactionRow] = []
        formula_rows = formula_sheet.iter_rows(
            min_row=TRANSACTION_FIRST_ROW,
            max_col=len(TRANSACTION_HEADERS),
            values_only=True,
        )
        value_rows = value_sheet.iter_rows(
            min_row=TRANSACTION_FIRST_ROW,
            max_col=len(TRANSACTION_HEADERS),
            values_only=True,
        )
        for row_number, (entered_row, effective_row) in enumerate(
            zip_longest(formula_rows, value_rows, fillvalue=()),
            start=TRANSACTION_FIRST_ROW,
        ):
            entered = list(entered_row or ())[: len(TRANSACTION_HEADERS)]
            effective = list(effective_row or ())[: len(TRANSACTION_HEADERS)]
            entered.extend([None] * (len(TRANSACTION_HEADERS) - len(entered)))
            effective.extend([None] * (len(TRANSACTION_HEADERS) - len(effective)))
            if not any(value is not None for value in effective[:12]):
                continue

            values = {
                header: effective[index]
                for index, header in enumerate(TRANSACTION_HEADERS)
            }
            formulas = {
                header: value
                for header, value in zip(TRANSACTION_HEADERS, entered, strict=True)
                if isinstance(value, str) and value.startswith("=")
            }
            raw_cells = {
                header: {
                    "entered": _json_value(entered[index]),
                    "effective": _json_value(effective[index]),
                }
                for index, header in enumerate(TRANSACTION_HEADERS)
            }
            rows.append(
                WorkbookTransactionRow(
                    row_number=row_number,
                    values=values,
                    formulas=formulas,
                    raw_source_data={
                        "sheet": "Transactions",
                        "row_number": row_number,
                        "cells": raw_cells,
                    },
                )
            )
        return rows

    @staticmethod
    def _read_holdings(values_book: Any) -> dict[tuple[str, str], Decimal]:
        sheet = values_book["Holdings"]
        headers = tuple(sheet.cell(4, column).value for column in range(1, 5))
        if headers != ("Account", "Asset", "Target Bucket", "Quantity"):
            raise WorkbookStructureError(
                "Holdings reconciliation headers are unexpected"
            )

        holdings: dict[tuple[str, str], Decimal] = {}
        for row in sheet.iter_rows(min_row=5, max_col=4, values_only=True):
            values = list(row or ())[:4]
            values.extend([None] * (4 - len(values)))
            account = values[0]
            symbol = values[1]
            quantity = values[3]
            if account in (None, "") or symbol in (None, ""):
                continue
            holdings[(str(account).strip(), str(symbol).strip().upper())] = Decimal(
                str(quantity or 0)
            )
        return holdings
